import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace', line_buffering=True)
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace', line_buffering=True)

from dotenv import load_dotenv
load_dotenv()

import requests
import json
import time
import os
import shutil
import schedule
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from google import genai
import pdfplumber

# ============================================================
# CONFIGURATION
# ============================================================
GEMINI_KEYS = [
    os.environ.get("GEMINI_KEY_1", ""),
    os.environ.get("GEMINI_KEY_2", ""),
    os.environ.get("GEMINI_KEY_3", ""),
]
GEMINI_KEYS = [k for k in GEMINI_KEYS if k]  # Remove empty keys
gemini_key_index = 0
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
BASE_FOLDER = r"C:\Users\LENOVO\Desktop\linkdin projects"
SUBSCRIBERS_FILE = os.path.join(BASE_FOLDER, "subscribers.txt")
FILINGS_FOLDER = os.path.join(BASE_FOLDER, "Filings")
EXCEL_FILE = os.path.join(BASE_FOLDER, "NSE_Filings_Log.xlsx")
SEEN_FILINGS_FILE = os.path.join(BASE_FOLDER, "seen_filings.json")
WATCHLIST_FILE = os.path.join(BASE_FOLDER, "watchlist.json")

# Price movement tracking: {symbol: {"price": float, "time": datetime}}
price_tracker = {}

# ============================================================
# GEMINI SETUP
# ============================================================
gemini_client = genai.Client(api_key=GEMINI_KEYS[0])

def rotate_gemini_key():
    global gemini_key_index, gemini_client
    gemini_key_index = (gemini_key_index + 1) % len(GEMINI_KEYS)
    gemini_client = genai.Client(api_key=GEMINI_KEYS[gemini_key_index])
    print(f"   Rotated to Gemini key #{gemini_key_index + 1}")

# ============================================================
# STYLES
# ============================================================
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HIGH_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
MODERATE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
COLUMN_WIDTHS = [12, 10, 12, 25, 30, 12, 40, 10, 12, 40, 40, 12, 10]
HEADERS = [
    "Date", "Time", "Symbol", "Company", "Filing Type", "Category",
    "Summary", "Verdict", "Confidence", "Reason", "Risk",
    "CMP at Filing", "Day Change %"
]

# ============================================================
# GLOBAL NSE SESSION
# ============================================================
nse_session = None

# ============================================================
# FOLDER + EXCEL SETUP
# ============================================================
def setup():
    os.makedirs(FILINGS_FOLDER, exist_ok=True)
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NSE Filings"
        ws.append(HEADERS)
        for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        wb.save(EXCEL_FILE)
        print("   ✅ Excel file created.")

# ============================================================
# CLASSIFY FILINGS
# ============================================================
def classify_filing(subject, pdf_text=""):
    combined = (subject + " " + pdf_text[:500]).lower()
    high_impact = [
        "acquisition", "merger", "demerger", "amalgamation", "scheme of arrangement",
        "scheme", "fund rais", "fundrais", "joint venture", "product launch",
        "new order", "new project", "recognition", "disinvestment", "diversification",
        "operational update", "commencement", "commercial production", "agreement",
        "partnership", "nvidia", "wins", "awarded", "secures", "bags",
        "signs mou", "launches", "expands", "enters agreement", "strategic",
        "order win", "contract win", "letter of intent", "loi", "mou signed",
        "jv", "takeover", "buyback", "delisting"
    ]
    moderate = [
        "board meeting", "results", "financial results", "press release",
        "general updates", "update", "credit rating", "change in management",
        "appointment", "resignation", "insider trading", "bulk deal", "block deal",
        "investor presentation", "analysts meet", "institutional investor",
        "con. call", "trading window", "allotment"
    ]
    for k in high_impact:
        if k in combined:
            return "HIGH"
    for k in moderate:
        if k in combined:
            return "MODERATE"
    return "ROUTINE"

# ============================================================
# SEEN FILINGS
# ============================================================
def load_seen():
    if os.path.exists(SEEN_FILINGS_FILE):
        with open(SEEN_FILINGS_FILE, "r") as f:
            return set(json.load(f))
    return set()

def save_seen(seen):
    with open(SEEN_FILINGS_FILE, "w") as f:
        json.dump(list(seen), f)

# ============================================================
# WATCHLIST
# ============================================================
def load_watchlist():
    if os.path.exists(WATCHLIST_FILE):
        with open(WATCHLIST_FILE, "r") as f:
            return json.load(f)
    return {}

def is_watchlisted(symbol):
    watchlist = load_watchlist()
    for group, symbols in watchlist.items():
        if symbol in symbols:
            return True, group
    return False, None

# ============================================================
# PRICE MOVEMENT TRACKER
# ============================================================
def track_price(symbol, price):
    """Store baseline CMP at time of filing for later comparison."""
    try:
        base_price = float(str(price).replace(",", ""))
        price_tracker[symbol] = {
            "price": base_price,
            "time": datetime.now(),
        }
        print(f"   Tracking {symbol} baseline: Rs {base_price}")
    except (ValueError, TypeError):
        pass

def check_price_movements():
    """Check if tracked HIGH filing stocks moved >2% within 10 min."""
    if not price_tracker:
        return
    session = get_session()
    now = datetime.now()
    expired = []
    for symbol, data in list(price_tracker.items()):
        elapsed = (now - data["time"]).total_seconds()
        if elapsed < 600:  # Not yet 10 minutes
            continue
        expired.append(symbol)
        try:
            current_price, _ = get_live_price(session, symbol)
            if current_price == "N/A":
                continue
            current = float(str(current_price).replace(",", ""))
            baseline = data["price"]
            if baseline <= 0:
                continue
            pct_change = ((current - baseline) / baseline) * 100
            if abs(pct_change) >= 2.0:
                direction = "+" if pct_change > 0 else ""
                print(f"   PRICE ALERT: {symbol} moved {direction}{pct_change:.1f}% since filing!")
                send_price_alert(symbol, baseline, current, pct_change)
        except Exception as e:
            print(f"   Price check error for {symbol}: {e}")
    for sym in expired:
        del price_tracker[sym]

def send_price_alert(symbol, baseline, current, pct_change):
    """Send follow-up Telegram alert for significant price movement."""
    direction = "+" if pct_change > 0 else ""
    arrow = "\U0001f4c8" if pct_change > 0 else "\U0001f4c9"
    message = f"""{arrow} PRICE MOVING: {symbol} {direction}{pct_change:.1f}%
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501
\U0001f4b0 CMP at Filing: Rs {baseline:.2f}
\U0001f4b0 Current CMP: Rs {current:.2f}
\U0001f4ca Movement: {direction}{pct_change:.1f}%
\u23f0 Within 10 min of filing
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501
\u26a1 Post-filing price action detected"""

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    chat_ids = load_subscribers()
    for chat_id in chat_ids:
        try:
            requests.post(url, json={"chat_id": chat_id, "text": message}, timeout=10)
        except Exception as e:
            print(f"   Price alert Telegram error: {e}")

# ============================================================
# TELEGRAM SUBSCRIBERS
# ============================================================
def load_subscribers():
    if not os.path.exists(SUBSCRIBERS_FILE):
        with open(SUBSCRIBERS_FILE, "w") as f:
            f.write("1281388903\n")
    with open(SUBSCRIBERS_FILE, "r") as f:
        ids = [line.strip() for line in f.readlines() if line.strip()]
    return ids

# ============================================================
# NSE SESSION
# ============================================================
def create_nse_session():
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Connection": "keep-alive"
    })
    try:
        print("   Warming up NSE session...")
        session.get("https://www.nseindia.com", timeout=15)
        time.sleep(5)
        session.get("https://www.nseindia.com/market-data/live-equity-market", timeout=10)
        time.sleep(5)
        session.headers.update({
            "Accept": "application/json, text/plain, */*",
            "Referer": "https://www.nseindia.com/market-data/live-equity-market",
            "X-Requested-With": "XMLHttpRequest"
        })
        print("   ✅ NSE session ready.")
    except Exception as e:
        print(f"   Session setup warning: {e}")
    return session

def get_session():
    global nse_session
    if nse_session is None:
        nse_session = create_nse_session()
    return nse_session

def reset_session():
    global nse_session
    print("   Resetting NSE session...")
    nse_session = None

# ============================================================
# FETCH FILINGS (3 PARALLEL NSE CALLS)
# ============================================================
def fetch_filings(session):
    today = datetime.now().strftime("%d-%m-%Y")
    urls = {
        "equities": f"https://www.nseindia.com/api/corporate-announcements?index=equities&from_date={today}&to_date={today}",
        "sme": f"https://www.nseindia.com/api/corporate-announcements?index=sme&from_date={today}&to_date={today}",
        "debt": f"https://www.nseindia.com/api/corporate-announcements?index=debt&from_date={today}&to_date={today}",
    }
    headers = dict(session.headers)
    headers.pop("Accept-Encoding", None)

    counts = {}

    def fetch_one(name_url):
        name, url = name_url
        try:
            response = session.get(url, timeout=15, headers=headers)
            if response.status_code == 200 and len(response.text) > 100:
                data = response.json()
                counts[name] = len(data)
                return data
        except Exception as e:
            print(f"   Error fetching {name}: {e}")
        counts[name] = 0
        return []

    combined = []
    seen_ids = set()
    with ThreadPoolExecutor(max_workers=3) as pool:
        results = list(pool.map(fetch_one, urls.items()))
    for filings in results:
        for f in filings:
            fid = f"{f.get('symbol', '')}_{f.get('seqNo', '')}_{f.get('an_dt', '')}"
            if fid not in seen_ids:
                seen_ids.add(fid)
                combined.append(f)

    eq = counts.get("equities", 0)
    sm = counts.get("sme", 0)
    dt = counts.get("debt", 0)
    print(f"   ✅ Fetched {len(combined)} total filings for today (equities: {eq}, sme: {sm}, debt: {dt})")
    return combined

# ============================================================
# GET LIVE STOCK PRICE
# ============================================================
def get_live_price(session, symbol):
    try:
        url = f"https://www.nseindia.com/api/quote-equity?symbol={symbol}"
        response = session.get(url, timeout=10)
        if response.status_code == 200:
            data = response.json()
            price_info = data.get("priceInfo", {})
            price = price_info.get("lastPrice", "N/A")
            change_pct = price_info.get("pChange", "N/A")
            return price, change_pct
    except Exception as e:
        print(f"   Price fetch error for {symbol}: {e}")
    return "N/A", "N/A"

# ============================================================
# DOWNLOAD PDF
# ============================================================
def download_pdf(session, filing, symbol):
    try:
        attach_url = filing.get("attchmntFile", "") or filing.get("attachment", "")
        if not attach_url:
            return None
        if not attach_url.startswith("http"):
            attach_url = "https://www.nseindia.com" + attach_url

        company_folder = os.path.join(FILINGS_FOLDER, symbol)
        os.makedirs(company_folder, exist_ok=True)

        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        subject_clean = (filing.get("desc", "") or filing.get("subject", "filing"))[:40]
        subject_clean = "".join(c for c in subject_clean if c.isalnum() or c in " -_").strip()
        filename = f"{date_str}_{subject_clean}.pdf"
        filepath = os.path.join(company_folder, filename)

        pdf_session = requests.Session()
        pdf_session.headers.update(session.headers)
        pdf_session.cookies.update(session.cookies)
        response = pdf_session.get(attach_url, timeout=20)
        if response.status_code == 200:
            with open(filepath, "wb") as f:
                f.write(response.content)
            print(f"   PDF saved: {filename}")
            return filepath
    except Exception as e:
        print(f"   PDF download error: {e}")
    return None

# ============================================================
# EXTRACT TEXT FROM PDF
# ============================================================
def extract_pdf_text(pdf_path):
    try:
        text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:5]:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text[:4000]
    except Exception as e:
        print(f"   PDF text extraction error: {e}")
    return ""

# ============================================================
# GEMINI ANALYSIS (HIGH FILINGS ONLY)
# ============================================================
def analyze_with_gemini(filing_text, subject, symbol, price, change_pct):
    prompt = f"""You are a senior Indian equity analyst. Analyze this NSE corporate filing and give a sharp market impact assessment.

COMPANY: {symbol}
FILING TYPE: {subject}
CURRENT PRICE: Rs {price}
TODAY'S CHANGE: {change_pct}%

FILING CONTENT:
{filing_text if filing_text else "[Full PDF not available — assess based on filing type and subject only]"}

ASSESSMENT RULES:
- Acquisition: Bullish if strategic fit, Bearish if overpriced or debt-funded
- Fund Raising (QIP/Rights/NCD): Mildly Bearish (dilution) unless strong growth story
- Merger/Demerger: Usually Bullish if value unlocking, check terms
- Joint Venture: Bullish if strong partner or new market
- New Order/Project: Bullish — bigger order = stronger signal
- Product Launch: Bullish if large addressable market
- Business/Operational Update: Depends on content — read carefully
- Disinvestment: Bullish if monetizing non-core assets
- If stock already up more than 5% today — likely priced in, reduce confidence
- If stock is flat today — genuine surprise, increase confidence

Respond in EXACTLY this format, no extra text:
SUMMARY: [3 lines max, plain English, what the company actually announced]
VERDICT: [BULLISH or BEARISH or NEUTRAL]
CONFIDENCE: [HIGH or MEDIUM or LOW]
REASON: [One sentence why]
RISK: [One sentence — what could go wrong or what to watch]"""

    for attempt in range(len(GEMINI_KEYS)):
        try:
            response = gemini_client.models.generate_content(
                model="gemini-2.0-flash-lite",
                contents=prompt
            )
            return response.text
        except Exception as e:
            error_str = str(e)
            if "429" in error_str and attempt < len(GEMINI_KEYS) - 1:
                print(f"   Gemini key #{gemini_key_index + 1} rate limited. Rotating...")
                rotate_gemini_key()
            else:
                print(f"   Gemini error: {e}")
                break
    return f"SUMMARY: {subject}\nVERDICT: NEUTRAL\nCONFIDENCE: LOW\nREASON: Analysis unavailable\nRISK: Review manually"

# ============================================================
# PARSE GEMINI RESPONSE
# ============================================================
def parse_gemini(text):
    result = {
        "summary": "See filing",
        "verdict": "NEUTRAL",
        "confidence": "LOW",
        "reason": "N/A",
        "risk": "Review manually"
    }
    for line in text.strip().split("\n"):
        if line.startswith("SUMMARY:"):
            result["summary"] = line.replace("SUMMARY:", "").strip()
        elif line.startswith("VERDICT:"):
            result["verdict"] = line.replace("VERDICT:", "").strip().upper()
        elif line.startswith("CONFIDENCE:"):
            result["confidence"] = line.replace("CONFIDENCE:", "").strip().upper()
        elif line.startswith("REASON:"):
            result["reason"] = line.replace("REASON:", "").strip()
        elif line.startswith("RISK:"):
            result["risk"] = line.replace("RISK:", "").strip()
    return result

# ============================================================
# SEND TELEGRAM
# ============================================================
def send_telegram(filing, symbol, company, analysis, price, change_pct):
    verdict = analysis["verdict"]
    if "BULLISH" in verdict:
        verdict_emoji = "BULLISH \U0001f7e2"
    elif "BEARISH" in verdict:
        verdict_emoji = "BEARISH \U0001f534"
    else:
        verdict_emoji = "NEUTRAL \U0001f7e1"

    conf = analysis["confidence"]
    conf_emoji = "\u2705" if conf == "HIGH" else "\u26a1" if conf == "MEDIUM" else "\u26a0\ufe0f"

    message = f"""\U0001f514 NSE FILING ALERT
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501
\U0001f3e2 {symbol} | {company}
\U0001f4cb {filing.get('desc', '') or filing.get('subject', 'Corporate Filing')}
\u23f0 {datetime.now().strftime('%H:%M IST | %d %b %Y')}
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501
\U0001f4dd SUMMARY
{analysis['summary']}
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501
\U0001f4ca MARKET IMPACT
Verdict: {verdict_emoji}
Confidence: {conf} {conf_emoji}
Reason: {analysis['reason']}
\u26a0\ufe0f Risk: {analysis['risk']}
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501
\U0001f4c8 PRICE AT FILING
CMP: Rs {price} | Change: {change_pct}%
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501"""

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    chat_ids = load_subscribers()
    for chat_id in chat_ids:
        try:
            resp = requests.post(url, json={"chat_id": chat_id, "text": message}, timeout=10)
            if resp.status_code == 200:
                print(f"   \u2705 Telegram sent to {chat_id}")
            else:
                print(f"   Telegram error for {chat_id}: {resp.text}")
        except Exception as e:
            print(f"   Telegram exception for {chat_id}: {e}")

# ============================================================
# SEND WATCHLIST TELEGRAM ALERT
# ============================================================
def send_watchlist_telegram(filing, symbol, company, category, group, price, change_pct):
    message = f"""\u2b50 WATCHLIST ALERT: {symbol}
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501
\U0001f3e2 {symbol} | {company}
\U0001f4c1 Watchlist: {group}
\U0001f4cb {filing.get('desc', '') or filing.get('subject', 'Corporate Filing')}
\U0001f3f7 Category: {category}
\u23f0 {datetime.now().strftime('%H:%M IST | %d %b %Y')}
\U0001f4b0 CMP: Rs {price} | Change: {change_pct}%
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501
\u26a1 Your watchlisted stock filed a corporate announcement"""

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    chat_ids = load_subscribers()
    for chat_id in chat_ids:
        try:
            requests.post(url, json={"chat_id": chat_id, "text": message}, timeout=10)
        except Exception as e:
            print(f"   Watchlist Telegram error: {e}")

# ============================================================
# LOG TO EXCEL
# ============================================================
def log_to_excel(filing, symbol, company, analysis, price, change_pct, category):
    now = datetime.now()
    row_data = [
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        symbol,
        company,
        filing.get("desc", "") or filing.get("subject", ""),
        category,
        analysis["summary"],
        analysis["verdict"],
        analysis["confidence"],
        analysis["reason"],
        analysis["risk"],
        str(price),
        f"{float(change_pct):.2f}%" if change_pct != "N/A" else "N/A"
    ]

    for attempt in range(3):
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.append(row_data)
            row_idx = ws.max_row
            if category == "HIGH":
                for cell in ws[row_idx]:
                    cell.fill = HIGH_FILL
            elif category == "MODERATE":
                for cell in ws[row_idx]:
                    cell.fill = MODERATE_FILL
            temp_file = EXCEL_FILE.replace(".xlsx", "_temp.xlsx")
            backup_file = EXCEL_FILE.replace(".xlsx", "_backup.xlsx")
            wb.save(temp_file)
            # Keep a backup before replacing
            if os.path.exists(EXCEL_FILE):
                shutil.copy2(EXCEL_FILE, backup_file)
            shutil.move(temp_file, EXCEL_FILE)
            print(f"   \u2705 Excel logged: {symbol} [{category}]")
            return
        except PermissionError:
            if attempt < 2:
                print(f"   Excel locked, retrying in 2s... ({attempt + 1}/3)")
                time.sleep(2)
            else:
                print(f"   \u26a0\ufe0f Excel locked — skipping log for {symbol}. Close the file!")
        except Exception as e:
            print(f"   Excel error: {e}")
            return

# ============================================================
# MAIN CHECK FUNCTION
# ============================================================
def check_filings():
    print(f"\n{'=' * 50}")
    print(f"\U0001f50d NSE Check at {datetime.now().strftime('%H:%M:%S on %d %b %Y')}")
    print(f"{'=' * 50}")

    seen = load_seen()
    session = get_session()
    filings = fetch_filings(session)

    filings = [f for f in filings if f is not None]

    if not filings:
        print("\u26a0\ufe0f  No data from NSE. Resetting session for next cycle.")
        reset_session()
        return

    new_count = 0

    for filing in filings:
        symbol = filing.get("symbol", "").upper().strip()
        subject = filing.get("desc", "") or filing.get("subject", "")
        seq = filing.get("seqNo", "")
        an_dt = filing.get("an_dt", "")
        filing_id = f"{symbol}_{seq}_{an_dt}"

        if filing_id in seen:
            continue

        new_count += 1
        seen.add(filing_id)

        # Step 1: Quick classify by subject only
        category = classify_filing(subject)

        # Step 2: Get live price
        price, change_pct = get_live_price(session, symbol)

        # Step 3: Download PDF
        pdf_path = download_pdf(session, filing, symbol)
        company = filing.get("corp_name", symbol)

        # Step 4: For HIGH/MODERATE — extract PDF text and reclassify
        pdf_text = ""
        if category in ("HIGH", "MODERATE") and pdf_path:
            pdf_text = extract_pdf_text(pdf_path)
            category = classify_filing(subject, pdf_text)

        # Step 5: Check watchlist
        watched, watch_group = is_watchlisted(symbol)
        priority_tag = f" [WATCHLIST: {watch_group}]" if watched else ""

        print(f"\n\U0001f195 {symbol}: {subject} [{category}]{priority_tag}")
        print(f"   Rs {price} | {change_pct}%")

        # Step 6: Gemini analysis only for HIGH
        if category == "HIGH":
            print(f"   Analyzing with Gemini...")
            raw = analyze_with_gemini(pdf_text, subject, symbol, price, change_pct)
            analysis = parse_gemini(raw)
            print(f"   Verdict: {analysis['verdict']} | Confidence: {analysis['confidence']}")
            send_telegram(filing, symbol, company, analysis, price, change_pct)
            # Track price for movement alert
            track_price(symbol, price)
        else:
            analysis = {
                "summary": "\u2014",
                "verdict": "N/A",
                "confidence": "N/A",
                "reason": "Routine filing - no analysis needed",
                "risk": "N/A"
            }

        # Step 7: Watchlist alert (triggers for ALL categories)
        if watched and category != "HIGH":  # HIGH already sent via send_telegram
            print(f"   Sending watchlist alert for {symbol}...")
            send_watchlist_telegram(filing, symbol, company, category, watch_group, price, change_pct)

        # Step 8: Log ALL to Excel
        log_to_excel(filing, symbol, company, analysis, price, change_pct, category)

        time.sleep(2)

    save_seen(seen)

    # Check price movements for tracked HIGH filings
    check_price_movements()

    if new_count == 0:
        print("\u2705 No new filings this cycle.")
    else:
        print(f"\n\u2705 {new_count} new filing(s) processed.")

# ============================================================
# STARTUP TELEGRAM MESSAGE
# ============================================================
def send_startup_message():
    message = f"""\U0001f7e2 NSE MONITOR STARTED
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501
\u23f0 {datetime.now().strftime('%H:%M IST | %d %b %Y')}
\U0001f4e1 Watching: ALL NSE companies
\U0001f3af High Impact Alerts: Telegram
\U0001f4ca All filings: Excel log
\u23f1 Interval: Every 30 seconds
\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501"""
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    chat_ids = load_subscribers()
    for chat_id in chat_ids:
        try:
            requests.post(url, json={"chat_id": chat_id, "text": message}, timeout=10)
        except Exception as e:
            print(f"   Startup message error for {chat_id}: {e}")

# ============================================================
# ENTRY POINT
# ============================================================
if __name__ == "__main__":
    print("\U0001f680 NSE Corporate Filings Monitor")
    print("   Tracking: ALL NSE companies (equities + SME + debt)")
    print("   Alerts: HIGH impact \u2192 Telegram | ALL \u2192 Excel")
    print("   Interval: Every 30 seconds\n")

    setup()

    # Verify Telegram connectivity before starting
    print("   Testing Telegram connection...")
    test_url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    test_ids = load_subscribers()
    for cid in test_ids:
        try:
            r = requests.post(test_url, json={
                "chat_id": cid,
                "text": f"\u2705 Monitor started successfully\n\u23f0 {datetime.now().strftime('%H:%M IST | %d %b %Y')}\n\U0001f4e1 Telegram connection verified"
            }, timeout=10)
            if r.status_code == 200:
                print(f"   \u2705 Telegram test OK (chat {cid})")
            else:
                print(f"   \u26a0\ufe0f Telegram test FAILED for {cid}: {r.text}")
        except Exception as e:
            print(f"   \u26a0\ufe0f Telegram test FAILED: {e}")

    print("   Initializing NSE session...")
    nse_session = create_nse_session()
    send_startup_message()
    check_filings()

    schedule.every(30).seconds.do(check_filings)

    print("\n\u23f0 Monitor running. Keep this terminal open during market hours.")
    print("   Press Ctrl+C to stop.\n")

    while True:
        schedule.run_pending()
        time.sleep(10)
